# -*- coding: utf-8 -*-
"""경쟁사 이벤트 크롤링 (GitHub / 로컬). Playwright + SQLite + 엑셀 출력."""
import sys
import warnings
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from playwright.sync_api import sync_playwright

import event_db
from config import ALERT_PREFIX, CORP, COLUMNS, EXCEL_PREFIX, OUTPUT_DIR
from crawler import crawl_all

warnings.filterwarnings('ignore')


def buho(text):
    if not isinstance(text, str):
        return text
    return (text.replace('&#36;', '$').replace('&#37;', '%').replace('&#38;', '&')
            .replace('&#162;', '¢').replace('&#163;', '£').replace('&#165;', '¥')
            .replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&')
            .replace('&quot;', '"').replace('&#35;', '#').replace('&#39;', "'"))


def excel_cont(rows_df, ws):
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 70
    ws.column_dimensions['F'].width = 70
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13

    for r in dataframe_to_rows(rows_df, index=False, header=True):
        ws.append(r)

    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            if column_cells[0].column_letter == 'D':
                cell.value = '=HYPERLINK("{}", "{}")'.format(cell.value, '링크') if cell.value != '' else ''
                cell.font = Font(size=9, italic=True, underline='singleAccounting', color='0000ff')
            if column_cells[0].column_letter in ['A', 'B', 'C', 'D', 'G', 'H']:
                cell.alignment = Alignment(horizontal='center')

    for cell in ws['1']:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, bold=True)
        cell.border = Border(bottom=Side(border_style='thin'))
        green = PatternFill(start_color='d8e4bc', end_color='d8e4bc', patternType='solid')
        cell.fill = green


def run_crawl():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print('[crawl] Playwright 시작...')
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=['--disable-blink-features=AutomationControlled'],
        )
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1400, 'height': 900},
            locale='ko-KR',
        )
        context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined});")
        page = context.new_page()
        try:
            rows, namu, stats, errors = crawl_all(page)
        finally:
            browser.close()

    rows_df = pd.DataFrame(rows, columns=COLUMNS)
    if not namu.empty:
        rows_all = pd.concat([rows_df, namu], ignore_index=True)
    else:
        rows_all = rows_df.copy()
    rows_all['제목'] = rows_all['제목'].apply(buho)

    print('[증권사별 수집]')
    fail = 0
    for corp in CORP:
        val = stats.get(corp, rows_all[rows_all['증권사'] == corp].shape[0])
        if isinstance(val, str):
            fail += 1
            print(f'  {corp}: 오류 - {errors.get(corp, val)}')
        else:
            print(f'  {corp}: {val}건')
    print(f'  합계: {len(rows_all)}건')
    return rows_all, fail


def build_outputs(rows_all):
    today = datetime.today().strftime('%Y%m%d')
    data_all = event_db.load_yesterday()

    wb = Workbook()
    for j, corp in enumerate(CORP):
        data = data_all[data_all['증권사'] == corp]
        rows_cor = rows_all[rows_all['증권사'] == corp]
        wb.create_sheet(corp, j)
        ws = wb[corp]
        excel_cont(rows_cor, ws)

        new = set(rows_cor['제목']) - set(data['제목'])
        if new:
            for title in new:
                for row_index in range(2, ws.max_row + 1):
                    subj = ws.cell(row=row_index, column=5).value
                    if subj == title:
                        ws.cell(row=row_index, column=5).font = Font(size=9, bold=True, color='00FF0000')

    data = event_db.load_yesterday()
    ws = wb['Sheet']
    ws.title = '전체'
    excel_cont(rows_all, ws)

    rows_all = rows_all.copy()
    rows_all['key'] = rows_all['증권사'] + '[]' + rows_all['제목']
    data['key'] = data['증권사'] + '[]' + data['제목'].astype(str)

    new = sorted(set(rows_all['key']) - set(data['key']))
    if new:
        for key in new:
            corp_name, subj = key.split('[]', 1)
            for row_index in range(2, ws.max_row + 1):
                name = ws.cell(row=row_index, column=1).value
                title = ws.cell(row=row_index, column=5).value
                if name == corp_name and title == subj:
                    ws.cell(row=row_index, column=5).font = Font(size=9, bold=True, color='00FF0000')

        result = []
        for key in new:
            temp = '- ' + key.replace('[]', ' / ').replace('"', '') + '\n'
            if key not in 'NH투자증권':
                temp = temp + ' ' + rows_all[rows_all['key'] == key]['url'].values[0] + '\n'
            result.append(temp)

        alert_path = OUTPUT_DIR / f'{ALERT_PREFIX}{today}.txt'
        result.insert(0, '경쟁사 신규 이벤트가 ' + str(len(new)) + '건 업로드 되었습니다. \n')
        alert_path.write_text('\n'.join(map(str, result)), encoding='utf-8')
        print(f'[알림] 신규 {len(new)}건 -> {alert_path}')
    else:
        print('[알림] 신규 이벤트 없음')

    xlsx_path = OUTPUT_DIR / f'{EXCEL_PREFIX}{today}.xlsx'
    wb.save(xlsx_path)
    print(f'[엑셀] {xlsx_path}')

    event_db.save_snapshot(today, rows_all[COLUMNS])
    print(f'[DB] {event_db.DB_PATH}')
    return len(new)


def main():
    print('=' * 60)
    print('경쟁사 이벤트 크롤링 (GitHub)')
    print(f'output: {OUTPUT_DIR}')
    print('=' * 60)

    rows_all, fail = run_crawl()
    build_outputs(rows_all)

    print('=' * 60)
    if fail:
        print(f'완료 (오류 증권사 {fail}개)')
        sys.exit(1)
    print('완료')
    sys.exit(0)


if __name__ == '__main__':
    main()
